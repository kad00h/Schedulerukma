import openpyxl

def basics(path):
    work_book = openpyxl.load_workbook(path)
    work_sheet = work_book.active
    rows_amount = work_sheet.max_row
    columns_amount = work_sheet.max_column

    #Отримуємо доступ ддо таблиці і знахдимо кількість заповнених клітинок по горизонталі і вертикалі щоб ними оперувати в циклах

    subject_learning = []
    pair_time = []
    weeks = []
    place = []
    group = []
    weekday_lst = ["понеділок", "вівторок", "середа", "четвер", "п'ятниця", "субота"]

    # Робимо списки для того щоб туди додавати предмети і створюємо словник щоб потім з нього виводити данні через гет валю

    for subj in range(2, rows_amount):
        subject_learning.append(work_sheet.cell(row = subj, column=3).value)
    for wek in range(2, rows_amount):
        weeks.append(work_sheet.cell(row = wek, column=5).value)
    for par in range(2, rows_amount):
        pair_time.append(work_sheet.cell(row = par, column=2).value)
    for pl in range(2, rows_amount):
        place.append(work_sheet.cell(row = pl, column=6).value)
    for gr in range(2, rows_amount):
        group.append(work_sheet.cell(row = gr, column=4).value)
    print("ФІ \n    ІПЗ \n"  )
    a = 0
    for i in range(rows_amount):
        print("         " + str(group[i]))
        print("             " + str(subject_learning[i]))
        print("                 " + str(pair_time[i]))
        print("                     " + str(weeks[i]))
        print("                         " + str(place[i]))
        if a % 7 == 0:
            print("                             " + weekday_lst[a])

basics("3.xlsx")
